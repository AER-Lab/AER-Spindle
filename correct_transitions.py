import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side



def correct_transitions(file_path):
    """
    Correct transitions in the W-R Transitions sheet and highlight corrections in orange, 
    and transitions in yellow.
    """
    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name="W-R Transitions")

    # Initialize counters and details list
    w_r_transitions = 0
    r_epochs_after_w_r = 0
    w_epochs_before_w_r = 0
    transition_details = []

    # Add correction and notes columns if not present
    if 'Correction' not in df.columns:
        df['Correction'] = df['Prediction']
    if 'Notes' not in df.columns:
        df['Notes'] = ''
    else:
        df['Notes'] = df['Notes'].fillna('')

    # Apply rules for W-R transitions and corrections
    for i in range(1, len(df) - 1):
        if df.iloc[i - 1]['Prediction'] == 'W' and df.iloc[i]['Prediction'] == 'R':
            w_r_transitions += 1

            # Count R epochs after W-R transition
            r_count = 0
            for j in range(i, len(df)):
                if df.iloc[j]['Prediction'] == 'R':
                    r_epochs_after_w_r += 1
                    r_count += 1
                else:
                    break

            # Count W epochs before W-R transition
            w_count = 0
            for k in range(i - 1, -1, -1):
                if df.iloc[k]['Prediction'] == 'W':
                    w_epochs_before_w_r += 1
                    w_count += 1
                else:
                    break

            # Store the details of the transition
            transition_details.append({
                'Transition Index': i,
                'W Epochs Before': w_count,
                'R Epochs After': r_count
            })

            # Add notes and corrections
            if pd.isna(df.at[i, 'Notes']):
                df.at[i, 'Notes'] = ''
            df.at[i, 'Notes'] += f'W-R transition at index {i}, W before: {w_count}, R after: {r_count}'

            # Rule 1: Correct R to W if W > 4 before and either W/NR follows R
            if w_count > 4 and (df.iloc[i + 1]['Prediction'] in ['W', 'NR'] or df.iloc[i + 2]['Prediction'] in ['W', 'NR']):
                df.at[i, 'Correction'] = 'W'
                df.at[i, 'Notes'] += ' | Rule 1: Corrected, R changed to W due to >4 W before and W/NR after'

            # Rule 2: Handle W (1-4) before R with long NR preceding W
            if (
                1 <= w_count <= 4 and
                df.iloc[max(0, k - 5):k]['Prediction'].tolist().count('NR') > 5 and
                r_count > 5
            ):
                df.loc[max(0, k - w_count):k, 'Correction'] = 'NR'
                df.at[i, 'Notes'] += ' | Rule 2: Corrected, W changed to NR due to long NR before W and >5 R after'

    # Rule 3: Single R correction (NR >3 before, W after R)
    for i in range(1, len(df) - 4):
        if (
            df.iloc[i - 3:i]['Prediction'].tolist().count('NR') >= 3 and
            df.iloc[i]['Prediction'] == 'R' and
            df.iloc[i + 1:i + 5]['Prediction'].tolist().count('W') >= 3
        ):
            df.at[i, 'Correction'] = 'W'
            df.at[i, 'Notes'] += ' | Rule 3: Corrected, Single R changed to W due to NR >3 before and W >3 after'

    # Rule 4: Single R correction with following R in 4 epochs
    for i in range(1, len(df) - 4):
        if (
            df.iloc[i - 3:i]['Prediction'].tolist().count('NR') >= 3 and
            df.iloc[i]['Prediction'] == 'R' and
            df.iloc[i + 1:i + 5]['Prediction'].tolist().count('R') > 2
        ):
            df.loc[i + 1:i + 5, 'Correction'] = 'R'
            df.loc[i + 1:i + 5, 'Notes'] += ' | Rule 4: Corrected, Following W changed to R due to >2 R after'

    # Rule 5: NR-R(single)-NR transition
    for i in range(1, len(df) - 1):
        if (
            df.iloc[i - 1]['Prediction'] == 'NR' and
            df.iloc[i]['Prediction'] == 'R' and
            df.iloc[i + 1]['Prediction'] == 'NR'
        ):
            df.at[i, 'Correction'] = 'NR'
            df.at[i, 'Notes'] += ' | Rule 5: Corrected, Single R changed to NR'

    # Rule 6: Single NR correction (W-NR(single)-W)
    for i in range(1, len(df) - 1):
        if (
            df.iloc[i - 1]['Prediction'] == 'W' and
            df.iloc[i]['Prediction'] == 'NR' and
            df.iloc[i + 1]['Prediction'] == 'W'
        ):
            df.at[i, 'Correction'] = 'W'
            df.at[i, 'Notes'] += ' | Rule 6: Corrected, Single NR changed to W due to W before and after'

    # Rule 7: Long R (>4) - NR(single) - W (>3)
    for i in range(6, len(df) - 4):
        if (
            df.iloc[i - 4:i]['Prediction'].tolist().count('R') > 4 and
            df.iloc[i]['Prediction'] == 'NR' and
            df.iloc[i + 1:i + 4]['Prediction'].tolist().count('W') > 3
        ):
            df.at[i, 'Correction'] = 'W'
            df.at[i, 'Notes'] += ' | Rule 7: Corrected, NR changed to W due to >4 R before and >3 W after'

    # Rule 8: Long R (>6) to NR (>6)
    for i in range(12, len(df)):
        if (
            df.iloc[i - 6:i]['Prediction'].tolist().count('R') > 6 and
            df.iloc[i:i + 6]['Prediction'].tolist().count('NR') > 6
        ):
            df.loc[i - 2:i, 'Correction'] = 'W'
            df.loc[i - 2:i, 'Notes'] += ' | Rule 8: Corrected, Artificial W inserted for long R to NR'

    # Rule 9: Correct last state in the document
    if df.iloc[-1]['Prediction'] in ['R', 'NR']:
        df.at[len(df) - 1, 'Correction'] = 'W'
        df.at[len(df) - 1, 'Notes'] += ' | Rule 9: Corrected last state to W'

    # Rule 10: Check if last 3 were W, and at the signal W-R-R, turn into W
    for i in range(2, len(df)):
        if (
            df.iloc[i - 2]['Prediction'] == 'W' and
            df.iloc[i - 1]['Prediction'] == 'R' and
            df.iloc[i]['Prediction'] == 'R'
        ):
            df.at[i - 1, 'Correction'] = 'W'
            df.at[i, 'Correction'] = 'W'
            df.at[i - 1, 'Notes'] += ' | Rule 10: Corrected, R-R changed to W-W due to W before'
            df.at[i, 'Notes'] += ' | Rule 10: Corrected, R-R changed to W-W due to W before'
    # Rule 11: If W > 4 and W-R signal and multiple R > 3 and W > 3 after, turn R to W
    for i in range(1, len(df) - 4):
        if (
            df.iloc[i - 4:i]['Prediction'].tolist().count('W') > 4 and
            df.iloc[i]['Prediction'] == 'R' and
            df.iloc[i + 1:i + 4]['Prediction'].tolist().count('R') > 3 and
            df.iloc[i + 1:i + 4]['Prediction'].tolist().count('W') > 3
        ):
            df.at[i, 'Correction'] = 'W'
            df.at[i, 'Notes'] += ' | Rule 11: Corrected, R changed to W due to >4 W before, >3 R and >3 W after'
    # Save corrections directly to the W-R Transitions sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, sheet_name='W-R Transitions')

    # Apply formatting
    wb = load_workbook(file_path)
    ws = wb['W-R Transitions']
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        transition_value = row[3].value  # Assuming 'Transition Highlight' is the 4th column
        notes_value = row[5].value  # Assuming 'Notes' is the 6th column
        if notes_value and 'Corrected' in notes_value:  # Highlight corrections in orange
            for cell in row:
                cell.fill = orange_fill
        elif transition_value == 'Yes':  # Highlight W-R transitions in yellow
            for cell in row:
                cell.fill = yellow_fill
        for cell in row:  # Apply thin borders to all rows
            cell.border = thin_border
    
    

    wb.save(file_path)

    return w_r_transitions, r_epochs_after_w_r, w_epochs_before_w_r, transition_details
# Example usage
file_path = r'C:\Users\geosaad\Desktop\Main-Scripts\SpindleModelWeights_compare\Spindle-Prediction-Compare\Model_Comparison\Validate_State_Transitions\pre-1341_predictions_W-R_Transitions.xlsx'
w_r_transitions, r_epochs_after_w_r, w_epochs_before_w_r, transition_details = correct_transitions(file_path)
print(f"Number of W-R transitions: {w_r_transitions}")
print(f"Number of R epochs after W-R transitions: {r_epochs_after_w_r}")
print(f"Number of W epochs before W-R transitions: {w_epochs_before_w_r}")

# Print detailed information for each transition
for detail in transition_details:
    print(f"Transition Index: {detail['Transition Index']}, W Epochs Before: {detail['W Epochs Before']}, R Epochs After: {detail['R Epochs After']}")