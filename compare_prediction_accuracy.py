import os
import glob
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import ttest_ind
import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows




def compute_confusion_matrix(data, labels):
    # Assuming `data` is the predictions and `labels` are the actual labels
    return pd.crosstab(labels, data, normalize='index') * 100

def plot_confusion_matrix(conf_matrix, title):
    plt.figure(figsize=(10, 8))
    sns.heatmap(conf_matrix, annot=True, fmt=".2f", cmap="Blues", cbar=True)
    plt.title(title)
    plt.xlabel('Predicted Label')
    plt.ylabel('True Label')
    plt.show()

def plot_class_accuracy(class_accuracy, title):
    # Create the bar plot
    ax = class_accuracy.plot(kind='bar', color='skyblue', figsize=(10, 10))
    plt.title(title)
    plt.xlabel('Class')
    plt.ylabel('Accuracy (%)')
    plt.xticks(rotation=45)
    plt.ylim(0, 100)

    # Annotate the bars with the percentage values
    for p in ax.patches:  # Loop over every bar
        ax.annotate(f"{p.get_height():.2f}%",  # This formats the number as a percentage
                    (p.get_x() + p.get_width() / 2., p.get_height()),  # This positions the text in the center of the bar
                    ha='center',  # Horizontally center the text
                    va='center',  # Vertically center the text
                    xytext=(0, 10),  # Position text 10 points above the top of the bar
                    textcoords='offset points')  # Use offset points to position the text

    plt.show()





def compare_predictions_and_labels(prediction_file, label_file, writer, sheet_name):
    # Load the CSV files
    predictions = pd.read_csv(prediction_file)
    labels = pd.read_csv(label_file, header=None, names=['Time', 'Label'])
    prediction_file_name = os.path.basename(prediction_file)
    label_file_name = os.path.basename(label_file)
    labels['Label'] = labels['Label'].replace(regex={r'W.*': 'W', r'R.*': 'R', r'NR.*': 'NR'})

    print("Comparing {} and {}".format(prediction_file_name, label_file_name))
    
    # Combine predictions and labels into a single DataFrame for easier analysis
    combined = pd.DataFrame({'Prediction': predictions['State'], 'Label': labels['Label']})
    combined['Correct'] = combined['Prediction'] == combined['Label']
    
    # Calculate overall accuracy
    overall_accuracy = combined['Correct'].mean() * 100
    
    # Calculate class-specific accuracy
    class_accuracy = combined.groupby('Label')['Correct'].mean() * 100  

    # Create both percentage and frequency matrices
    misclassification_matrix_pct = pd.crosstab(combined['Label'], combined['Prediction'], normalize='index') * 100
    misclassification_matrix_freq = pd.crosstab(combined['Label'], combined['Prediction'])

    # Extract values for NR, R, and W from percentage matrix
    nr_value = misclassification_matrix_pct.at['NR', 'NR'] if 'NR' in misclassification_matrix_pct.index and 'NR' in misclassification_matrix_pct.columns else 0
    r_value = misclassification_matrix_pct.at['R', 'R'] if 'R' in misclassification_matrix_pct.index and 'R' in misclassification_matrix_pct.columns else 0
    w_value = misclassification_matrix_pct.at['W', 'W'] if 'W' in misclassification_matrix_pct.index and 'W' in misclassification_matrix_pct.columns else 0

    return overall_accuracy, class_accuracy, misclassification_matrix_pct, misclassification_matrix_freq, nr_value, r_value, w_value

def define_folder_path(folder_path):
    folder_path = folder_path
    excel_path = os.path.join(folder_path, "model_comparison_results.xlsx")
    prediction_files = glob.glob(os.path.join(folder_path, '*_predictions-correct.csv'))
    output_excel_path = os.path.join(folder_path, "mismatches_summary.xlsx")
    return folder_path, excel_path, prediction_files, output_excel_path


def compare_files(folder_path):
    folder_path, excel_path, prediction_files, output_excel_path = define_folder_path(folder_path)

    print("Prediction files are: ", prediction_files)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        average_across_files = []
        combined_confusion_matrix = None
        for prediction_file in prediction_files:
            print("Processing file: {}".format(prediction_file))
            base_name = os.path.splitext(os.path.basename(prediction_file))[0].replace('_predictions-correct', '')
            print("Basename: ", base_name)
            label_file = os.path.join(folder_path, base_name + '.csv')
            print("prediction file and label file: ", prediction_file, label_file)
            if os.path.exists(label_file):
                sheet_name = base_name  # Sheet name based on file base name
                overall_accuracy, class_accuracy, misclassification_matrix, misclassification_matrix_freq, nr_value, r_value, w_value = compare_predictions_and_labels(prediction_file, label_file, writer, sheet_name)
                
                print("Confusion matrix for {}: \n{}".format(sheet_name, misclassification_matrix))

                if combined_confusion_matrix is None:
                    combined_confusion_matrix = misclassification_matrix
                else:
                    combined_confusion_matrix += misclassification_matrix

                plot_confusion_matrix(misclassification_matrix, f"Confusion Matrix for {sheet_name}")
                plot_confusion_matrix(misclassification_matrix_freq, f"Confusion Matrix for {sheet_name} (Frequency)")
                plot_class_accuracy(class_accuracy, f"Class Accuracy for {sheet_name}")

                average_across_files.append(overall_accuracy)
                print("Overall Accuracy: {:.2f}%".format(overall_accuracy))
                print("Class Accuracy: {}".format(class_accuracy))    
                
                # Create detailed results DataFrame
                results = pd.DataFrame({
                    'Metric': ['Overall Accuracy', 'NR Accuracy', 'R Accuracy', 'W Accuracy', 
                             'NR Total Samples', 'R Total Samples', 'W Total Samples'],
                    'Value': [
                        f"{overall_accuracy:.2f}%",
                        f"{nr_value:.2f}%",
                        f"{r_value:.2f}%",
                        f"{w_value:.2f}%",
                        misclassification_matrix_freq.sum(axis=1)['NR'] if 'NR' in misclassification_matrix_freq.index else 0,
                        misclassification_matrix_freq.sum(axis=1)['R'] if 'R' in misclassification_matrix_freq.index else 0,
                        misclassification_matrix_freq.sum(axis=1)['W'] if 'W' in misclassification_matrix_freq.index else 0
                    ]
                })

                # Write results to Excel
                results.to_excel(writer, sheet_name=sheet_name, startrow=0, index=False)
                
                # Add percentage confusion matrix
                writer.sheets[sheet_name].cell(row=10, column=1, value="Confusion Matrix (Percentages)")
                misclassification_matrix.to_excel(writer, sheet_name=sheet_name, startrow=11)
                
                # Add frequency confusion matrix
                writer.sheets[sheet_name].cell(row=17, column=1, value="Confusion Matrix (Frequencies)")
                misclassification_matrix_freq.to_excel(writer, sheet_name=sheet_name, startrow=18)
        if combined_confusion_matrix is not None:
            # Normalize the combined confusion matrix again to ensure it's a proper percentage
            combined_confusion_matrix = combined_confusion_matrix.div(combined_confusion_matrix.sum(axis=1), axis=0) * 100
            plot_confusion_matrix(combined_confusion_matrix, "Combined Confusion Matrix Across All Files")

        if average_across_files:
            average_accuracy = sum(average_across_files) / len(average_across_files)
            print(f"Average Accuracy across all files: {average_accuracy:.2f}%")
            
            # Write this average to a summary sheet in the Excel workbook
            pd.DataFrame({'Average Accuracy': [average_accuracy]}).to_excel(writer, sheet_name='Summary', index=False)
            
            # Optionally plot the average accuracy as a simple bar chart
            plt.figure(figsize=(6, 4))
            plt.bar("Average Accuracy", average_accuracy, color='skyblue')
            plt.title("Average Accuracy Across All Files")
            plt.ylabel("Accuracy (%)")
            plt.ylim(0, 100)
            plt.show()
    loop_files_to_compare(folder_path, output_excel_path)

def plot_mismatches(prediction_file, label_file):
    """
    Plots mismatches between prediction and label files in 2-hour segments.
    """
    print(f"Comparing {os.path.basename(prediction_file)} with {os.path.basename(label_file)}")

    # Load the CSV files
    predictions = pd.read_csv(prediction_file)
    labels = pd.read_csv(label_file, header=None, names=['Time', 'Label'])
    
    # Normalize label states
    labels['Label'] = labels['Label'].replace(regex={r'W.*': 'W', r'R.*': 'R', r'NR.*': 'NR'})

    # Combine predictions and labels
    combined = pd.DataFrame({'Prediction': predictions['State'], 'Label': labels['Label']})
    combined['Correct'] = combined['Prediction'] == combined['Label']

    # Determine the number of plots needed
    num_rows = len(combined)
    rows_per_plot = 2 * 3600 // 4  # 2 hours of data, assuming each row is 4 seconds
    num_plots = (num_rows // rows_per_plot) + 1

    for i in range(num_plots):
        start_row = i * rows_per_plot
        end_row = min((i + 1) * rows_per_plot, num_rows)

        # Plot the predictions and labels for the current segment
        plt.figure(figsize=(12, 6))
        plt.plot(combined['Prediction'][start_row:end_row].reset_index(drop=True), label='Prediction', color='blue')
        plt.plot(combined['Label'][start_row:end_row].reset_index(drop=True), label='Label', color='green')

        plt.title(f"Predictions vs Labels (Segment {i + 1})")
        plt.xlabel("Time (2-hour segments)")
        plt.ylabel("Sleep Stage")
        plt.legend()
        plt.show()


def save_mismatches_to_excel(prediction_file, label_file):
    """
    Identifies mismatches between prediction and label files, 
    calculates quarterly errors, and returns mismatch data.
    """
    # Load the CSV files
    predictions = pd.read_csv(prediction_file)
    labels = pd.read_csv(label_file, header=None, names=['Time', 'Label'])

    # Normalize label states
    labels['Label'] = labels['Label'].replace(regex={r'W.*': 'W', r'R.*': 'R', r'NR.*': 'NR'})
    base_name = os.path.splitext(os.path.basename(prediction_file))[0].replace('_predictions', '')

    # Combine predictions and labels
    combined = pd.DataFrame({'Prediction': predictions['State'], 'Label': labels['Label']})
    combined['Correct'] = combined['Prediction'] == combined['Label']

    # Find mismatches
    mismatches = combined[combined['Correct'] == False].copy()
    
    # Calculate quarterly errors
    mismatches['Quarter'] = pd.qcut(mismatches.index, 4, labels=['Q1', 'Q2', 'Q3', 'Q4'])
    quarterly_errors = mismatches['Quarter'].value_counts().sort_index()

    return mismatches, quarterly_errors


def analyze_mismatches(prediction_file, label_file):
    """
    Analyzes mismatches and computes error statistics.
    """
    # Load CSV files
    predictions = pd.read_csv(prediction_file)
    labels = pd.read_csv(label_file, header=None, names=['Time', 'Label'])

    # Normalize label states
    labels['Label'] = labels['Label'].replace(regex={r'W.*': 'W', r'R.*': 'R', r'NR.*': 'NR'})
    base_name = os.path.splitext(os.path.basename(prediction_file))[0].replace('_predictions-correct', '')

    # Combine predictions and labels
    combined = pd.DataFrame({'Prediction': predictions['State'], 'Label': labels['Label']})
    combined['Correct'] = combined['Prediction'] == combined['Label']

    # Calculate statistics
    mismatches = combined[combined['Correct'] == False]
    nr_errors = mismatches[mismatches['Label'] == 'NR'].shape[0]
    r_errors = mismatches[mismatches['Label'] == 'R'].shape[0]
    w_errors = mismatches[mismatches['Label'] == 'W'].shape[0]
    total_errors = len(mismatches)
    accuracy = (len(combined) - total_errors) / len(combined) * 100

    # Create summary row
    summary = pd.DataFrame({
        'Filename': [base_name],
        'Total Errors': [total_errors],
        'NR Errors': [nr_errors],
        'R Errors': [r_errors],
        'W Errors': [w_errors],
        'Accuracy': [accuracy]
    })

    return summary


def write_summary_to_excel(all_mismatches, all_quarterly_errors, all_summaries, output_file):
    """
    Writes all mismatches, quarterly errors, and summary statistics to an Excel file.
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for filename, mismatches in all_mismatches.items():
            mismatches.to_excel(writer, sheet_name=f'{filename}_mismatch', index=True)

        # Write quarterly errors summary
        quarterly_summary = pd.DataFrame.from_dict(all_quarterly_errors, orient='index').fillna(0)
        quarterly_summary.columns = ['Q1', 'Q2', 'Q3', 'Q4']
        quarterly_summary.to_excel(writer, sheet_name='Quarterly_Errors')

        # Write summary statistics
        summary_df = pd.concat(all_summaries, ignore_index=True)
        summary_df.to_excel(writer, sheet_name='Summary_Statistics', index=False)

    # Apply formatting
    wb = load_workbook(output_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:  # Bold headers
            cell.font = Font(bold=True)
        
        # Add borders
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Highlight mismatched 'R' in Prediction column (Red)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Highlight mismatched 'R' in Label column (Yellow)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row in ws.iter_rows(min_row=2):
            if row[1].value == 'R':  # Prediction column
                row[1].fill = red_fill
            if row[2].value == 'R':  # Label column
                row[2].fill = yellow_fill

    wb.save(output_file)


def loop_files_to_compare(folder_path, output_excel_path):
    """
    Loops through files in the folder, compares prediction and label files, 
    and writes mismatch data to an Excel file.
    """
    csv_files = glob.glob(os.path.join(folder_path, '*.csv'))
    prediction_files = glob.glob(os.path.join(folder_path, '*_predictions-correct.csv'))

    all_mismatches = {}
    all_quarterly_errors = {}
    all_summaries = []

    for label_file in csv_files:
        prediction_file = label_file.replace('.csv', '_predictions-correct.csv')
        if prediction_file in prediction_files:
            base_name = os.path.splitext(os.path.basename(prediction_file))[0].replace('_predictions-correct', '')
            plot_mismatches(prediction_file, label_file)
            mismatches, quarterly_errors = save_mismatches_to_excel(prediction_file, label_file)
            summary = analyze_mismatches(prediction_file, label_file)

            all_mismatches[base_name] = mismatches
            all_quarterly_errors[base_name] = quarterly_errors.to_dict()
            all_summaries.append(summary)

    # Write results to Excel
    write_summary_to_excel(all_mismatches, all_quarterly_errors, all_summaries, output_excel_path)